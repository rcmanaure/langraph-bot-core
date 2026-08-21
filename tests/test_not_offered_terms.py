"""Not-offered terms: admin CRUD + Excel template/upload (#53) + the
TTL-cached loading service (app/services/not_offered_terms.py). Modeled
directly on tests/test_canned_answers.py — same seam, same fixture shapes —
per the spec's Testing Decisions (this table mirrors canned_answers minus
the answer column)."""
import io
import sys
import types
from unittest.mock import AsyncMock, MagicMock, patch

for _mod in (
    "pypdf", "filetype", "tiktoken",
    "langchain_openai", "langgraph", "langgraph.graph",
    "langchain_core.vectorstores", "langchain_core.embeddings",
):
    if _mod not in sys.modules:
        sys.modules[_mod] = types.ModuleType(_mod)

if "app.services.indexer" not in sys.modules:
    _indexer_stub = types.ModuleType("app.services.indexer")
    async def _stub_run_index_job(*args, **kwargs):  # noqa: E301
        pass
    _indexer_stub.run_index_job = _stub_run_index_job
    sys.modules["app.services.indexer"] = _indexer_stub

import openpyxl  # noqa: E402
import pytest  # noqa: E402
from fastapi import FastAPI  # noqa: E402
from httpx import ASGITransport, AsyncClient  # noqa: E402

from app.auth import verify_operator_key  # noqa: E402
from app.models.not_offered_term import NotOfferedTerm  # noqa: E402
from app.routes.admin import router  # noqa: E402
from app.services import not_offered_terms  # noqa: E402


def make_app() -> FastAPI:
    app = FastAPI()
    app.include_router(router)
    app.dependency_overrides[verify_operator_key] = lambda: None
    return app


async def req(app, method, path, **kwargs):
    async with AsyncClient(transport=ASGITransport(app=app), base_url="http://test") as c:
        return await c.request(method.upper(), path, **kwargs)


def _session(**overrides):
    s = AsyncMock()
    s.execute = overrides.get("execute", AsyncMock())
    s.scalar = overrides.get("scalar", AsyncMock(return_value=1))
    s.commit = AsyncMock()
    s.refresh = overrides.get("refresh", AsyncMock())
    s.add = MagicMock()
    return s


def _ctx(session):
    ctx = AsyncMock()
    ctx.__aenter__ = AsyncMock(return_value=session)
    ctx.__aexit__ = AsyncMock(return_value=None)
    return ctx


def _scalars_result(rows):
    r = MagicMock()
    r.scalars.return_value.all.return_value = rows
    return r


def _scalar_one_result(row):
    r = MagicMock()
    r.scalar_one_or_none.return_value = row
    return r


def _xlsx_bytes(rows: list[str], header: str = "header") -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([header])
    for row in rows:
        ws.append([row])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ═════════════════════════════════════════════════════════════════════════════
# Admin CRUD
# ═════════════════════════════════════════════════════════════════════════════

@pytest.mark.asyncio
async def test_list_unknown_tenant_returns_404():
    app = make_app()
    sess = _session(scalar=AsyncMock(return_value=None))
    with patch("app.routes.admin.AsyncSessionLocal", return_value=_ctx(sess)):
        r = await req(app, "get", "/admin/tenants/ghost/not-offered-terms")
    assert r.status_code == 404


@pytest.mark.asyncio
async def test_list_returns_rows_for_tenant():
    app = make_app()
    row = NotOfferedTerm(id=1, tenant_id=1, keywords=["urocultivo"], match_mode="any")
    sess = _session(execute=AsyncMock(return_value=_scalars_result([row])))
    with patch("app.routes.admin.AsyncSessionLocal", return_value=_ctx(sess)):
        r = await req(app, "get", "/admin/tenants/acme/not-offered-terms")
    assert r.status_code == 200
    assert r.json() == [{"id": 1, "keywords": ["urocultivo"], "match_mode": "any"}]


@pytest.mark.asyncio
async def test_create_success_invalidates_cache():
    app = make_app()

    def _refresh(obj):
        obj.id = 1
    sess = _session(refresh=AsyncMock(side_effect=_refresh))

    with patch("app.routes.admin.AsyncSessionLocal", return_value=_ctx(sess)), \
         patch("app.services.not_offered_terms.invalidate") as mock_invalidate:
        r = await req(app, "post", "/admin/tenants/acme/not-offered-terms",
                      json={"keywords": ["urocultivo", "cultivo de orina"], "match_mode": "any"})
    assert r.status_code == 201
    body = r.json()
    assert body["keywords"] == ["urocultivo", "cultivo de orina"]
    mock_invalidate.assert_called_once_with("acme")


@pytest.mark.asyncio
async def test_create_empty_keywords_returns_422():
    app = make_app()
    r = await req(app, "post", "/admin/tenants/acme/not-offered-terms", json={"keywords": []})
    assert r.status_code == 422


@pytest.mark.asyncio
async def test_create_unknown_tenant_returns_404():
    app = make_app()
    sess = _session(scalar=AsyncMock(return_value=None))
    with patch("app.routes.admin.AsyncSessionLocal", return_value=_ctx(sess)):
        r = await req(app, "post", "/admin/tenants/ghost/not-offered-terms", json={"keywords": ["x"]})
    assert r.status_code == 404


@pytest.mark.asyncio
async def test_patch_updates_keywords_and_invalidates_cache():
    app = make_app()
    row = NotOfferedTerm(id=5, tenant_id=1, keywords=["urocultivo"], match_mode="any")
    sess = _session(execute=AsyncMock(return_value=_scalar_one_result(row)))

    with patch("app.routes.admin.AsyncSessionLocal", return_value=_ctx(sess)), \
         patch("app.services.not_offered_terms.invalidate") as mock_invalidate:
        r = await req(app, "patch", "/admin/tenants/acme/not-offered-terms/5",
                      json={"keywords": ["resonancia magnetica"]})
    assert r.status_code == 200
    assert r.json()["keywords"] == ["resonancia magnetica"]
    assert row.keywords == ["resonancia magnetica"]
    mock_invalidate.assert_called_once_with("acme")


@pytest.mark.asyncio
async def test_patch_not_found_returns_404():
    app = make_app()
    sess = _session(execute=AsyncMock(return_value=_scalar_one_result(None)))
    with patch("app.routes.admin.AsyncSessionLocal", return_value=_ctx(sess)):
        r = await req(app, "patch", "/admin/tenants/acme/not-offered-terms/999", json={"keywords": ["x"]})
    assert r.status_code == 404


@pytest.mark.asyncio
async def test_delete_success_invalidates_cache():
    app = make_app()
    delete_result = MagicMock()
    delete_result.rowcount = 1
    sess = _session(execute=AsyncMock(return_value=delete_result))

    with patch("app.routes.admin.AsyncSessionLocal", return_value=_ctx(sess)), \
         patch("app.services.not_offered_terms.invalidate") as mock_invalidate:
        r = await req(app, "delete", "/admin/tenants/acme/not-offered-terms/5")
    assert r.status_code == 204
    mock_invalidate.assert_called_once_with("acme")


@pytest.mark.asyncio
async def test_delete_not_found_returns_404():
    app = make_app()
    delete_result = MagicMock()
    delete_result.rowcount = 0
    sess = _session(execute=AsyncMock(return_value=delete_result))
    with patch("app.routes.admin.AsyncSessionLocal", return_value=_ctx(sess)):
        r = await req(app, "delete", "/admin/tenants/acme/not-offered-terms/999")
    assert r.status_code == 404


# ═════════════════════════════════════════════════════════════════════════════
# Excel template download + upload (#53)
# ═════════════════════════════════════════════════════════════════════════════

@pytest.mark.asyncio
async def test_download_template_returns_readable_xlsx_with_header_and_examples():
    app = make_app()
    r = await req(app, "get", "/admin/tenants/acme/not-offered-terms/template")
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]

    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    ws = wb.active
    rows = [row[0] for row in ws.iter_rows(min_col=1, max_col=1, values_only=True)]
    assert "no ofrece" in rows[0].lower()
    assert len(rows) > 1  # header + at least one example row
    # Example rows must be unmistakably marked and never real vocabulary --
    # an operator who re-uploads the template unedited must not end up
    # denying a real customer term (found in /code-review).
    for example_row in rows[1:]:
        assert "borre esta fila" in example_row.lower()


@pytest.mark.asyncio
async def test_upload_replaces_existing_list():
    app = make_app()
    xlsx = _xlsx_bytes(["urocultivo, cultivo de orina", "radiografia"])
    sess = _session()

    with patch("app.routes.admin.AsyncSessionLocal", return_value=_ctx(sess)), \
         patch("app.services.not_offered_terms.invalidate") as mock_invalidate:
        r = await req(
            app, "post", "/admin/tenants/acme/not-offered-terms/upload",
            files={"file": ("terms.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
    assert r.status_code == 200
    assert r.json() == {"imported": 2}
    mock_invalidate.assert_called_once_with("acme")
    # Delete (replace-all) happened before the new rows were added.
    delete_call = sess.execute.await_args_list[0]
    assert "DELETE FROM not_offered_terms" in str(delete_call.args[0])
    assert sess.add.call_count == 2


@pytest.mark.asyncio
async def test_upload_skips_blank_rows():
    app = make_app()
    xlsx = _xlsx_bytes(["urocultivo", "", "   ", "radiografia"])
    sess = _session()

    with patch("app.routes.admin.AsyncSessionLocal", return_value=_ctx(sess)), \
         patch("app.services.not_offered_terms.invalidate"):
        r = await req(
            app, "post", "/admin/tenants/acme/not-offered-terms/upload",
            files={"file": ("terms.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
    assert r.status_code == 200
    assert r.json() == {"imported": 2}


@pytest.mark.asyncio
async def test_upload_with_zero_valid_rows_rejects_and_touches_no_db():
    """A well-formed .xlsx with no non-blank data past the header (e.g. the
    template re-uploaded unedited, or data in the wrong column) must not
    silently wipe the tenant's existing list (found in /code-review)."""
    app = make_app()
    xlsx = _xlsx_bytes([])  # header only, no data rows
    with patch("app.routes.admin.AsyncSessionLocal") as mock_session_local:
        r = await req(
            app, "post", "/admin/tenants/acme/not-offered-terms/upload",
            files={"file": ("terms.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
    assert r.status_code == 400
    mock_session_local.assert_not_called()


@pytest.mark.asyncio
async def test_upload_malformed_file_returns_400_and_touches_no_db():
    app = make_app()
    with patch("app.routes.admin.AsyncSessionLocal") as mock_session_local:
        r = await req(
            app, "post", "/admin/tenants/acme/not-offered-terms/upload",
            files={"file": ("terms.xlsx", b"not a real xlsx file", "application/octet-stream")},
        )
    assert r.status_code == 400
    mock_session_local.assert_not_called()


@pytest.mark.asyncio
async def test_upload_oversized_file_returns_400_and_touches_no_db():
    """No cap meant an operator-key holder could force a full in-memory
    buffer + openpyxl parse of an arbitrarily large/pathological file
    (found in /code-review; MAX_MEDIA_BYTES is the existing precedent for
    bounding an inbound upload's size before it's touched)."""
    from app.routes.admin import _NOT_OFFERED_UPLOAD_MAX_BYTES

    app = make_app()
    oversized = b"x" * (_NOT_OFFERED_UPLOAD_MAX_BYTES + 1)
    with patch("app.routes.admin.AsyncSessionLocal") as mock_session_local:
        r = await req(
            app, "post", "/admin/tenants/acme/not-offered-terms/upload",
            files={"file": ("terms.xlsx", oversized, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
    assert r.status_code == 400
    mock_session_local.assert_not_called()


@pytest.mark.asyncio
async def test_upload_unknown_tenant_returns_404():
    app = make_app()
    xlsx = _xlsx_bytes(["urocultivo"])
    sess = _session(scalar=AsyncMock(return_value=None))
    with patch("app.routes.admin.AsyncSessionLocal", return_value=_ctx(sess)):
        r = await req(
            app, "post", "/admin/tenants/ghost/not-offered-terms/upload",
            files={"file": ("terms.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
    assert r.status_code == 404


# ═════════════════════════════════════════════════════════════════════════════
# app/services/not_offered_terms.py — TTL cache + invalidation
# ═════════════════════════════════════════════════════════════════════════════

@pytest.fixture(autouse=True)
def _clear_not_offered_terms_cache():
    not_offered_terms._cache.clear()
    yield
    not_offered_terms._cache.clear()


def _fetch_result(rows):
    r = MagicMock()
    r.fetchall.return_value = rows
    return r


def _row(id, keywords, match_mode):
    r = MagicMock()
    r.id, r.keywords, r.match_mode = id, keywords, match_mode
    return r


@pytest.mark.asyncio
async def test_get_not_offered_terms_caches_within_ttl():
    sess = _session(execute=AsyncMock(return_value=_fetch_result([_row(1, ["urocultivo"], "any")])))
    with patch("app.services.not_offered_terms.AsyncSessionLocal", return_value=_ctx(sess)):
        first = await not_offered_terms.get_not_offered_terms("acme")
        second = await not_offered_terms.get_not_offered_terms("acme")

    assert first == second == [{"id": 1, "keywords": ["urocultivo"], "match_mode": "any"}]
    sess.execute.assert_awaited_once()


@pytest.mark.asyncio
async def test_get_not_offered_terms_query_is_deterministically_ordered():
    sess = _session(execute=AsyncMock(return_value=_fetch_result([])))
    with patch("app.services.not_offered_terms.AsyncSessionLocal", return_value=_ctx(sess)):
        await not_offered_terms.get_not_offered_terms("acme")

    executed_sql = str(sess.execute.await_args.args[0])
    assert "ORDER BY" in executed_sql.upper()


@pytest.mark.asyncio
async def test_invalidate_forces_refetch_without_waiting_out_ttl():
    sess = _session(execute=AsyncMock(return_value=_fetch_result([_row(1, ["old"], "any")])))
    with patch("app.services.not_offered_terms.AsyncSessionLocal", return_value=_ctx(sess)):
        await not_offered_terms.get_not_offered_terms("acme")

    not_offered_terms.invalidate("acme")

    sess2 = _session(execute=AsyncMock(return_value=_fetch_result([_row(1, ["new"], "any")])))
    with patch("app.services.not_offered_terms.AsyncSessionLocal", return_value=_ctx(sess2)):
        result = await not_offered_terms.get_not_offered_terms("acme")

    assert result == [{"id": 1, "keywords": ["new"], "match_mode": "any"}]


@pytest.mark.asyncio
async def test_get_not_offered_terms_db_failure_degrades_to_empty_list():
    with patch("app.services.not_offered_terms.AsyncSessionLocal", side_effect=ConnectionError("db down")):
        result = await not_offered_terms.get_not_offered_terms("acme")
    assert result == []


# ═════════════════════════════════════════════════════════════════════════════
# app/services/not_offered_terms.py — match_not_offered_term (#53)
# ═════════════════════════════════════════════════════════════════════════════

@pytest.mark.asyncio
async def test_match_true_for_any_mode_single_keyword_hit():
    sess = _session(execute=AsyncMock(return_value=_fetch_result([_row(1, ["urocultivo"], "any")])))
    with patch("app.services.not_offered_terms.AsyncSessionLocal", return_value=_ctx(sess)):
        result = await not_offered_terms.match_not_offered_term("acme", "cuanto cuesta un urocultivo")
    assert result is True


@pytest.mark.asyncio
async def test_match_does_not_fire_on_a_keyword_embedded_in_another_word():
    sess = _session(execute=AsyncMock(return_value=_fetch_result([_row(1, ["rx"], "any")])))
    with patch("app.services.not_offered_terms.AsyncSessionLocal", return_value=_ctx(sess)):
        result = await not_offered_terms.match_not_offered_term("acme", "coordinacion administrativa")
    assert result is False


@pytest.mark.asyncio
async def test_match_ignores_accents_and_case():
    sess = _session(execute=AsyncMock(return_value=_fetch_result([_row(1, ["resonancia magnética"], "any")])))
    with patch("app.services.not_offered_terms.AsyncSessionLocal", return_value=_ctx(sess)):
        result = await not_offered_terms.match_not_offered_term("acme", "HACEN RESONANCIA MAGNETICA?")
    assert result is True


@pytest.mark.asyncio
async def test_match_all_mode_requires_every_keyword():
    sess = _session(execute=AsyncMock(return_value=_fetch_result([_row(1, ["cultivo", "orina"], "all")])))
    with patch("app.services.not_offered_terms.AsyncSessionLocal", return_value=_ctx(sess)):
        no_match = await not_offered_terms.match_not_offered_term("acme", "cultivo de sangre")
        full_match = await not_offered_terms.match_not_offered_term("acme", "cultivo de orina")
    assert no_match is False
    assert full_match is True


@pytest.mark.asyncio
async def test_match_returns_false_for_tenant_with_no_not_offered_terms():
    sess = _session(execute=AsyncMock(return_value=_fetch_result([])))
    with patch("app.services.not_offered_terms.AsyncSessionLocal", return_value=_ctx(sess)):
        result = await not_offered_terms.match_not_offered_term("acme", "cualquier cosa")
    assert result is False
